"""
excel_export.py -- Export job evaluations to an xlsx spreadsheet.

Provides inline updating of a jobs database Excel file during evaluation.
Each call to ``update_job_entry`` either appends a new row or overwrites an
existing row (deduplicated by URL).

Features:
  - Auto-creates the workbook + headers on first run
  - Bold, frozen header row with auto-fitted column widths
  - Score column is colour-coded (green / yellow / orange / red)
  - File-lock handling: retries once, then falls back to a timestamped copy
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column schema (order matters — indexes are 1-based in openpyxl)
# ---------------------------------------------------------------------------
COLUMNS: list[str] = [
    "#",
    "Company",
    "Title",
    "Score",
    "Recommendation",
    "German Required",
    "German Level",
    "Date Posted",
    "Summary",
    "Reasoning",
    "Skills Match",
    "Education Match",
    "Location Score",
    "Language Score",
    "Growth Score",
    "URL",
    "Evaluated At",
    "CV Generated",
]

# Column index helpers (1-based)
_COL = {name: idx + 1 for idx, name in enumerate(COLUMNS)}

# ---------------------------------------------------------------------------
# Score colour fills
# ---------------------------------------------------------------------------
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")

# Minimum / default column widths per header
_MIN_WIDTHS: dict[str, int] = {
    "#": 5,
    "Company": 22,
    "Title": 30,
    "Score": 8,
    "Recommendation": 16,
    "German Required": 16,
    "German Level": 14,
    "Date Posted": 14,
    "Summary": 50,
    "Reasoning": 50,
    "Skills Match": 13,
    "Education Match": 15,
    "Location Score": 14,
    "Language Score": 14,
    "Growth Score": 13,
    "URL": 40,
    "Evaluated At": 22,
    "CV Generated": 14,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def update_job_entry(
    evaluation: dict,
    output_path: str = "data/jobs_database.xlsx",
) -> str | None:
    """Add or update a single job row in the xlsx.

    Parameters
    ----------
    evaluation : dict
        The evaluation dict produced by ``Evaluator.evaluate()``.
    output_path : str
        Path to the xlsx workbook (created if missing).

    Returns
    -------
    str | None
        The path written to, or *None* if the write failed entirely.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # 1. Load or create workbook
    # ------------------------------------------------------------------
    wb, ws = _load_or_create(path)

    # ------------------------------------------------------------------
    # 2. Build row data from evaluation dict
    # ------------------------------------------------------------------
    scores = evaluation.get("scores", {})
    url = evaluation.get("url", "")

    row_data: dict[str, object] = {
        "Company": evaluation.get("company", evaluation.get("company_name", "")),
        "Title": evaluation.get("title", evaluation.get("job_title", "")),
        "Score": evaluation.get("global_score", ""),
        "Recommendation": evaluation.get("recommendation", ""),
        "German Required": "Yes" if evaluation.get("german_required") else "No",
        "German Level": evaluation.get("german_level_required", "unknown"),
        "Date Posted": evaluation.get("date_posted", ""),
        "Summary": evaluation.get("summary", ""),
        "Reasoning": evaluation.get("reasoning", ""),
        "Skills Match": scores.get("skills_match", ""),
        "Education Match": scores.get("education_match", ""),
        "Location Score": scores.get("location", ""),
        "Language Score": scores.get("language", ""),
        "Growth Score": scores.get("growth", ""),
        "URL": url,
        "Evaluated At": evaluation.get("evaluated_at", datetime.now().isoformat()),
        "CV Generated": "Yes" if evaluation.get("cv_path") else "No",
    }

    # ------------------------------------------------------------------
    # 3. Deduplicate by URL — find existing row or pick next free row
    # ------------------------------------------------------------------
    target_row = _find_row_by_url(ws, url) if url else None

    if target_row is not None:
        logger.info("Updating existing row %d for URL: %s", target_row, url)
        # Keep the original row number (#)
    else:
        target_row = ws.max_row + 1
        row_number = target_row - 1  # header is row 1
        row_data["#"] = row_number
        logger.info("Appending new row %d for: %s — %s",
                     target_row, row_data["Company"], row_data["Title"])

    # Write cells
    for col_name, value in row_data.items():
        col_idx = _COL[col_name]
        ws.cell(row=target_row, column=col_idx, value=value)

    # ------------------------------------------------------------------
    # 4. Apply score colour
    # ------------------------------------------------------------------
    _apply_score_fill(ws, target_row)

    # ------------------------------------------------------------------
    # 5. Auto-fit column widths & formatting pass
    # ------------------------------------------------------------------
    _autofit_columns(ws)

    # ------------------------------------------------------------------
    # 6. Save (with lock-safe fallback)
    # ------------------------------------------------------------------
    saved_path = _safe_save(wb, path)
    return saved_path


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _load_or_create(path: Path) -> tuple:
    """Return ``(Workbook, active sheet)``, creating headers if needed."""
    if path.exists():
        try:
            wb = load_workbook(path)
            ws = wb.active
            # Sanity-check: if headers are missing, re-write them
            if ws.cell(row=1, column=1).value != COLUMNS[0]:
                _write_headers(ws)
            return wb, ws
        except Exception as exc:
            logger.warning("Could not open existing workbook (%s); creating new.", exc)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    _write_headers(ws)
    return wb, ws


def _write_headers(ws) -> None:
    """Write header row with formatting + freeze."""
    for idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _find_row_by_url(ws, url: str) -> int | None:
    """Return the 1-based row index whose URL column matches *url*, or None."""
    if not url:
        return None
    url_col = _COL["URL"]
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=url_col).value
        if cell_value and str(cell_value).strip() == url.strip():
            return row_idx
    return None


def _apply_score_fill(ws, row: int) -> None:
    """Colour-code the Score cell for *row*."""
    score_col = _COL["Score"]
    cell = ws.cell(row=row, column=score_col)
    try:
        score = float(cell.value)
    except (TypeError, ValueError):
        return
    if score >= 4.0:
        cell.fill = FILL_GREEN
    elif score >= 3.5:
        cell.fill = FILL_YELLOW
    elif score >= 3.0:
        cell.fill = FILL_ORANGE
    else:
        cell.fill = FILL_RED


def _autofit_columns(ws) -> None:
    """Set each column width to ``max(min_width, longest_value + padding)``."""
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        min_w = _MIN_WIDTHS.get(col_name, 12)
        max_len = min_w
        letter = get_column_letter(col_idx)
        for cell in ws[letter]:
            if cell.value is not None:
                # Approximate: cap at 80 to avoid absurdly wide columns
                length = min(len(str(cell.value)) + 2, 80)
                if length > max_len:
                    max_len = length
        ws.column_dimensions[letter].width = max_len


def _safe_save(wb, path: Path) -> str | None:
    """Save the workbook, handling file-lock gracefully.

    Strategy:
    1. Try to save directly.
    2. On ``PermissionError`` (file open in Excel), try saving to a
       timestamped fallback and warn the user.
    """
    try:
        wb.save(str(path))
        logger.info("Excel saved: %s", path)
        return str(path)
    except PermissionError:
        # File is probably open in Excel — save a fallback copy
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = path.with_stem(f"{path.stem}_{ts}")
        try:
            wb.save(str(fallback))
            logger.warning(
                "Could not write to %s (file locked). Saved fallback: %s",
                path, fallback,
            )
            return str(fallback)
        except Exception as exc:
            logger.error("Failed to save Excel fallback: %s", exc)
            return None
    except Exception as exc:
        logger.error("Unexpected error saving Excel file: %s", exc)
        return None
